"""
Dataset-driven parser stress test for Spartan Shield Saver.

Usage:
    python test_ai_phrase_dataset_runner.py ai_phrase_test_dataset.xlsx
    python test_ai_phrase_dataset_runner.py ai_phrase_test_dataset.txt

Supported inputs:
- .xlsx : sheet named "cases" with columns:
    label | category | message | expected_types | state_last_event_id | notes
  expected_types should be pipe-separated, e.g. event_update|event_update_no_changes
- .txt / .csv / .tsv :
    expected_types<TAB>label<TAB>category<TAB>state_last_event_id<TAB>message

Output:
- ai_phrase_dataset_report.txt
"""

from pathlib import Path
from collections import Counter, defaultdict
import sys
import traceback

from ai.action_interpreter import interpret_message, get_default_chat_state

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


BASE_CONTEXT = {
    "events": [
        {
            "id": 1,
            "title": "Spring Expo",
            "date": "2026-04-24",
            "start_datetime": "2026-04-24T17:00:00",
            "location": "Webb Center Ballroom",
            "description": "Networking event",
            "guest_count": 80,
        },
        {
            "id": 2,
            "title": "Cybersecurity Mixer",
            "date": "2026-05-02",
            "start_datetime": "2026-05-02T18:30:00",
            "location": "Engineering Auditorium",
            "description": "Student mixer",
            "guest_count": 120,
        },
    ],
    "tasks": [
        {"id": 1, "event_id": 1, "title": "Call the caterer", "completed": 0},
        {"id": 2, "event_id": 1, "title": "Send invitations", "completed": 0},
        {"id": 3, "event_id": 1, "title": "Confirm venue", "completed": 0},
        {"id": 4, "event_id": 2, "title": "Prepare name tags", "completed": 0},
        {"id": 5, "event_id": 2, "title": "Review budget", "completed": 0},
    ],
}


def _split_expected(value):
    value = str(value or "").strip()
    if not value:
        return {"fallback"}
    return {piece.strip() for piece in value.split("|") if piece.strip()}


def load_cases_from_xlsx(path):
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx datasets.")
    wb = load_workbook(path, data_only=True)
    ws = wb["cases"] if "cases" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_map = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    for idx, header in enumerate(rows[0]):
        if header:
            header_map[str(header).strip().lower()] = idx

    required = ["label", "category", "message", "expected_types"]
    missing = [name for name in required if name not in header_map]
    if missing:
        raise RuntimeError(f"Missing required columns in dataset: {missing}")

    cases = []
    for row in rows[1:]:
        if not row or all(v in (None, "") for v in row):
            continue
        message = row[header_map["message"]]
        if not message:
            continue
        cases.append({
            "label": str(row[header_map["label"]]).strip(),
            "category": str(row[header_map["category"]]).strip(),
            "message": str(message).strip(),
            "expected_types": _split_expected(row[header_map["expected_types"]]),
            "state_last_event_id": int(row[header_map["state_last_event_id"]]) if "state_last_event_id" in header_map and row[header_map["state_last_event_id"]] not in (None, "") else None,
            "notes": str(row[header_map["notes"]]).strip() if "notes" in header_map and row[header_map["notes"]] is not None else "",
        })
    return cases


def load_cases_from_text(path):
    cases = []
    with open(path, "r", encoding="utf-8") as f:
        for line_no, raw in enumerate(f, 1):
            line = raw.rstrip("\n")
            if not line.strip() or line.lstrip().startswith("#"):
                continue
            parts = line.split("\t")
            if len(parts) < 5:
                raise RuntimeError(f"Line {line_no} needs 5 tab-separated fields: expected_types, label, category, state_last_event_id, message")
            expected, label, category, state_last_event_id, message = parts[:5]
            cases.append({
                "label": label.strip(),
                "category": category.strip(),
                "message": message.strip(),
                "expected_types": _split_expected(expected),
                "state_last_event_id": int(state_last_event_id) if state_last_event_id.strip() else None,
                "notes": "",
            })
    return cases


def load_cases(path_str):
    path = Path(path_str)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return load_cases_from_xlsx(path)
    if suffix in {".txt", ".tsv", ".csv"}:
        return load_cases_from_text(path)
    raise RuntimeError(f"Unsupported dataset format: {suffix}")


def run_case(case, context=None):
    context = context or BASE_CONTEXT
    state = get_default_chat_state()
    if case.get("state_last_event_id"):
        state["last_event_id"] = case["state_last_event_id"]
    try:
        result = interpret_message(case["message"], context=context, state=state)
        actual = result.get("type")
        status = "PASS" if actual in case["expected_types"] else "FAIL"
        return {
            **case,
            "status": status,
            "actual": actual,
            "result": result,
            "error": None,
        }
    except Exception as e:
        return {
            **case,
            "status": "ERROR",
            "actual": "ERROR",
            "result": None,
            "error": f"{type(e).__name__}: {e}",
            "traceback": traceback.format_exc(),
        }


def write_report(results, output_path="ai_phrase_dataset_report.txt"):
    counts = Counter(r["status"] for r in results)
    actual_counts = Counter(r["actual"] for r in results)
    by_category = defaultdict(list)
    for r in results:
        by_category[r["category"]].append(r)

    lines = []
    lines.append("AI Dataset Parser Report")
    lines.append("=" * 100)
    lines.append(f"Total cases: {len(results)}")
    lines.append(f"PASS: {counts.get('PASS', 0)}")
    lines.append(f"FAIL: {counts.get('FAIL', 0)}")
    lines.append(f"ERROR: {counts.get('ERROR', 0)}")
    lines.append("")
    lines.append("Returned action counts:")
    for k, v in actual_counts.most_common():
        lines.append(f"  - {k}: {v}")
    lines.append("")
    lines.append("Summary by category:")
    for category, items in sorted(by_category.items()):
        c = Counter(r["status"] for r in items)
        lines.append(f"  - {category}: PASS {c.get('PASS',0)}, FAIL {c.get('FAIL',0)}, ERROR {c.get('ERROR',0)}")
    lines.append("")

    failed = [r for r in results if r["status"] != "PASS"]
    if failed:
        lines.append("Failures / slips:")
        for r in failed:
            lines.append(f"[{r['status']}] {r['label']}")
            lines.append(f"  category: {r['category']}")
            lines.append(f"  input:    {r['message']}")
            lines.append(f"  expected: {sorted(r['expected_types'])}")
            lines.append(f"  actual:   {r['actual']}")
            if r.get("notes"):
                lines.append(f"  notes:    {r['notes']}")
            if r["error"]:
                lines.append(f"  error:    {r['error']}")
        lines.append("")

    lines.append("Detailed results:")
    for r in results:
        lines.append(f"{r['status']} | {r['label']}")
        lines.append(f"  category: {r['category']}")
        lines.append(f"  input:    {r['message']}")
        lines.append(f"  expected: {sorted(r['expected_types'])}")
        lines.append(f"  actual:   {r['actual']}")
        if r.get("notes"):
            lines.append(f"  notes:    {r['notes']}")
        if r["error"]:
            lines.append(f"  error:    {r['error']}")
        else:
            lines.append(f"  result:   {r['result']}")
        lines.append("")

    Path(output_path).write_text("\n".join(lines), encoding="utf-8")
    return output_path


def main():
    dataset_path = sys.argv[1] if len(sys.argv) > 1 else "ai_phrase_test_dataset.xlsx"
    cases = load_cases(dataset_path)
    if not cases:
        print("No cases found.")
        return
    results = [run_case(case) for case in cases]
    report = write_report(results)
    counts = Counter(r["status"] for r in results)
    print("=" * 100)
    print("AI DATASET PARSER RUN COMPLETE")
    print(f"Dataset: {dataset_path}")
    print(f"Total cases: {len(results)}")
    print(f"PASS: {counts.get('PASS', 0)}")
    print(f"FAIL: {counts.get('FAIL', 0)}")
    print(f"ERROR: {counts.get('ERROR', 0)}")
    print(f"Report written to: {report}")
    print("=" * 100)


if __name__ == "__main__":
    main()
